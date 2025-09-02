# -*- coding: utf-8 -*-
"""
Created on Tue Aug 26 15:19:48 2025

@author: Administrator
"""
import os
import sys
sys.path.append(os.getenv('TIA_SCRIPTING'))
import siemens_tia_scripting as STS
from openpyxl import Workbook
import shutil

Path_TIA_Project = 'Enter the file Path of Your TIA Portal Project Path, eg: C:\\TIA Portal Projet\\Sample.ap17'



# 1. Get Project Name - From the file Paths 
seperator = "\\"
positions = []
start = 0
while True:
    start = Path_TIA_Project.find(seperator, start)
    if start == -1:
        break
    positions.append(start)
    start += len(seperator)
Project_Name = Path_TIA_Project[positions[-1]+1:Path_TIA_Project.find('.ap')]

# 2. Define the Paths. (Siemens CAx file ends with *.aml)
base_dir = 'C:\\Users\\Administrator\\Documents\\01. Python Code\\15. TIA Scripting Python\\CAx\\'
aml_Path = base_dir+Project_Name+'.aml'
log_Path = base_dir+Project_Name+' Log file.log'

def RmvDir(path):   
    try:
        if os.path.exists(path):
            
            shutil.rmtree(path)
    except:
        pass

""" Objective:
    1. Read IP address from the CAx file
    2. Open the Hardware Configruation Page in TIA Portal
    3. Export the PLC Tags to an excel
"""

# 3. Open the TIA Portal
project = STS.open_attach_project(Path_TIA_Project)

# 3.1 Clean the contents
RmvDir('C:/Users/Administrator/Documents/01. Python Code/15. TIA Scripting Python/CAx')

# 3.2 Export the hardware config files
CAx_files = project.export_cax_data(export_file_path = aml_Path,log_file_path =log_Path )



# 4.0 Read IP addr from the CAx
import xml.etree.ElementTree as ET
tree = ET.parse(aml_Path)
root = tree.getroot()
# Convert to xml
tree.write('C:\\Users\\Administrator\\Documents\\01. Python Code\\15. TIA Scripting Python\\CAx\\'+Project_Name+'.xml', encoding='utf-8', xml_declaration=True)

""" 
Assumption: 
    
    We use IP address defined in PLC to go online. Therefore we need to find IP address allocated to the CPU.
    To do that, we need find the CPU reference, and hence find the ProfinetName Reference

"""

# >>> Objective 1: CPU's IP is the first element of Profinet Network. The element ID can be used to track where it is referenced inside of the CAx file <<<
CPU_ID_At_Left = root.find(".//InternalLink[@Name = 'Link To Subnet_1']").get('RefPartnerSideA').replace(':LogicalEndPoint_Node','')
CPU_element = root.findall(".//InternalElement/InternalElement")
IP_Online = []
for i in CPU_element:
    try:
        if CPU_ID_At_Left == i.get('ID'):
            main_PLC_IP_Online = i.find(".//Attribute[@Name='NetworkAddress']/Value").text
          
    except:
        pass

    

# >>> ------------- Objective 2: Open the Hardware Configruation Page in TIA Portal ------------- <<<

# Get Information from plcs
for plc in project.get_plcs():
    print(plc.get_name())
    plc.open_device_editor()
    properties = plc.get_properties()
    print(f'IP_Addr: {main_PLC_IP_Online}')
    # ------- Note here, tag_table is under plc. Refer to hierachy diagram ----------
    for tag_table in plc.get_plc_tag_tables():
        tag_table.export(target_directory_path = "C:\\Users\\Administrator\\Documents\\01. Python Code\\15. TIA Scripting Python\\export", export_options = STS.Enums.ExportOptions.WithDefaults
                         ,export_format = STS.Enums.ExportFormats.SimaticML)

        
#>>> Objective 3: Export PLC Tags and save to excel table
import XML_2_XLSX as Default_Tag_Table_xlsx
wb =  Workbook()
ws =  wb.active

# Title
Title_list = ['Name','Path','Data Type','Logical Address','Comment','Hmi Visible','Hmi Accessible','Hmi Writeable','Typeobject ID','Version ID']
for item in Title_list:
    ws.cell(1,Title_list.index(item)+1).value = item
# Write contents to the excel
for content in Default_Tag_Table_xlsx.Tag_name_elements:
    ws.cell(Default_Tag_Table_xlsx.Tag_name_elements.index(content)+2,1).value = content.text

for content1 in Default_Tag_Table_xlsx.logical_addr_elements:
    ws.cell(Default_Tag_Table_xlsx.logical_addr_elements.index(content1)+2,4).value = content1.text
    
for content2 in Default_Tag_Table_xlsx.Data_type_elements :
    ws.cell(Default_Tag_Table_xlsx.Data_type_elements.index(content2)+2,3).value = content2.text   
    
for content3 in Default_Tag_Table_xlsx.Comment_elements :
    ws.cell(Default_Tag_Table_xlsx.Comment_elements.index(content3)+2,5).value = content3.text   
    

# Rest , write default value to the columns
for i in range(0,len(Default_Tag_Table_xlsx.Tag_name_elements)):
    ws.cell(i+2,2).value = 'Default tag table'
    ws.cell(i+2,6).value = 'True'
    ws.cell(i+2,7).value = 'True'
    ws.cell(i+2,8).value = 'True'
New_Xlsx_path = Path_TIA_Project[:Path_TIA_Project.find('.ap')]    
wb.save(New_Xlsx_path+ '.xlsx')
os.startfile(New_Xlsx_path+ '.xlsx')


